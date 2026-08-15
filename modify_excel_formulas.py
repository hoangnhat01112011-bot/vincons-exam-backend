import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
import shutil
import os
import sys

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\DELL\Desktop\KPI_NhanCong_Vincons_Cleaned.xlsx"
BACKUP_PATH = r"C:\Users\DELL\Desktop\KPI_NhanCong_Vincons_Cleaned_Backup.xlsx"

def modify_excel():
    # 1. Back up the original file
    if os.path.exists(EXCEL_PATH):
        shutil.copy(EXCEL_PATH, BACKUP_PATH)
        print(f"Đã sao lưu file gốc sang: {BACKUP_PATH}")
    else:
        print(f"Lỗi: Không tìm thấy file Excel tại {EXCEL_PATH}")
        return

    # 2. Load workbook
    wb = openpyxl.load_workbook(EXCEL_PATH)
    
    # Define a helper function to set STT and formulas
    def apply_sheet_formulas(sheet_title):
        if sheet_title not in wb.sheetnames:
            return
        ws = wb[sheet_title]
        cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_indices = {name: idx for idx, name in enumerate(cols, 1) if name}
        
        # Helper to convert col number to letter (A, B, C, ...)
        def get_col_letter(col_idx):
            from openpyxl.utils import get_column_letter
            return get_column_letter(col_idx)

        for r in range(2, 200): # Write formulas for first 200 rows
            if 'STT' in col_indices:
                ws.cell(row=r, column=col_indices['STT'], value=r - 1)
                
            if sheet_title == 'GIAO VIỆC HÒA VỐN':
                # N: Đơn vị tính, O: Đơn giá NC sau +30%, P: Số ngày thi công, Q: Tổng quỹ lương/ngày của nhóm, R: Sản lượng hòa vốn 1 ngày, S: Sản lượng đạt hòa vốn / 1 ngày – Cả tổ, T: Tổng sản lượng hòa vốn cả đợt, U: Sản lượng theo khối lượng, V: Sản lượng theo thực tế thi công, W: Chênh lệch dự kiến Lãi(+)/Lỗ(-), X: ĐÁNH GIÁ & CẢNH BÁO
                ws[f'N{r}'] = f'=IF(F{r}="";"";IFERROR(VLOOKUP(F{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;3;FALSE);""))'
                ws[f'O{r}'] = f'=IF(F{r}="";"";IFERROR(VLOOKUP(F{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;5;FALSE);0))'
                ws[f'P{r}'] = f'=IF(OR(C{r}="";D{r}="");"";D{r}-C{r}+1)'
                ws[f'Q{r}'] = (
                    f'=I{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Tổ trưởng*") + '
                    f'J{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 1") + '
                    f'K{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 2") + '
                    f'L{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 3") + '
                    f'M{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ phụ")'
                )
                ws[f'R{r}'] = f'=IF(O{r}>0; Q{r}/O{r}; 0)'
                ws[f'S{r}'] = f'=IF(P{r}>0; H{r}/P{r}; 0)'
                ws[f'T{r}'] = f'=R{r}*P{r}'
                ws[f'U{r}'] = f'=H{r}*O{r}'
                ws[f'V{r}'] = f'=Q{r}*P{r}'
                ws[f'W{r}'] = f'=U{r}-V{r}'
                ws[f'X{r}'] = f'=IF(W{r}=0;"";IF(W{r}>0;"An toàn - Đạt định mức";"Cảnh báo - Không đạt định mức (LỖ)"))'
                
            elif sheet_title == 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY':
                # O: Đơn vị tính, P: Đơn giá NC sau +30%, Q: Số ngày thi công, R: Tổng quỹ lương/ngày của nhóm, S: Sản lượng theo thực tế thi công, T: Chênh lệch dự kiến Lãi(+)/Lỗ(-), U: ĐÁNH GIÁ & CẢNH BÁO
                ws[f'O{r}'] = f'=IF(G{r}="";"";IFERROR(VLOOKUP(G{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;3;FALSE);""))'
                ws[f'P{r}'] = f'=IF(G{r}="";"";IFERROR(VLOOKUP(G{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;5;FALSE);0))'
                ws[f'Q{r}'] = f'=IF(OR(D{r}="";E{r}="");"";E{r}-D{r}+1)'
                ws[f'R{r}'] = (
                    f'=J{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Tổ trưởng*") + '
                    f'K{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 1") + '
                    f'L{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 2") + '
                    f'M{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 3") + '
                    f'N{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ phụ")'
                )
                ws[f'S{r}'] = f'=R{r}*Q{r}'
                ws[f'T{r}'] = f'=(I{r}*P{r})-S{r}'
                ws[f'U{r}'] = f'=IF(T{r}=0;"";IF(T{r}>0;"An toàn - Đạt định mức";"Cảnh báo - Không đạt định mức (LỖ)"))'
                
            elif sheet_title == 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG':
                # F: Số ngày thi công, G: Tổng quỹ lương chi trả đến hiện tại của tổ, H: Tổng sản lượng theo khối lượng, I: Chênh lệch Lãi(+)/Lỗ(-), J: ĐÁNH GIÁ & CẢNH BÁO
                ws[f'F{r}'] = f'=IF(OR(D{r}="";E{r}="");"";E{r}-D{r}+1)'
                ws[f'G{r}'] = f'=SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"<>Tổ trưởng*")*F{r}'
                ws[f'H{r}'] = f'=SUMPRODUCT((\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$C$2:$C$5000=C{r})*(\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$I$2:$I$5000)*(\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$P$2:$P$5000))'
                ws[f'I{r}'] = f'=H{r}-G{r}'
                ws[f'J{r}'] = f'=IF(I{r}=0;"";IF(I{r}>0;"Lãi";"Lỗ"))'
                
            elif sheet_title == 'ĐÁNH GIÁ TỔNG DA':
                # E: Số ngày thi công, F: Tổng quỹ lương chi trả đến hiện tại, G: Tổng sản lượng theo khối lượng, H: Chênh lệch Lãi(+)/Lỗ(-), I: ĐÁNH GIÁ & CẢNH BÁO
                ws[f'E{r}'] = f'=IF(OR(C{r}="";D{r}="");"";D{r}-C{r}+1)'
                ws[f'F{r}'] = f'=SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"<>Tổ trưởng*")*E{r}'
                ws[f'G{r}'] = f'=SUMPRODUCT(\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$I$2:$I$5000*\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$P$2:$P$5000)'
                ws[f'H{r}'] = f'=G{r}-F{r}'
                ws[f'I{r}'] = f'=IF(H{r}=0;"";IF(H{r}>0;"Lãi";"Lỗ"))'

    # 3. Clear data validations (if any) to avoid duplicates
    ws_gvhv = wb['GIAO VIỆC HÒA VỐN']
    ws_gvhv.data_validations.dataValidation = []

    # 4. Set up Data Validations on GIAO VIỆC HÒA VỐN
    dv_to = DataValidation(type="list", formula1="='DANH MỤC TỔ'!$B$2:$B$150", allow_blank=True)
    dv_to.error ='Dữ liệu nhập không hợp lệ'
    dv_to.errorTitle = 'Lỗi nhập liệu'
    dv_to.prompt = 'Chọn Tổ thi công từ danh sách'
    dv_to.promptTitle = 'Tổ thi công'
    ws_gvhv.add_data_validation(dv_to)
    dv_to.add("B2:B100")

    dv_cap1 = DataValidation(type="list", formula1="='DANH MỤC HẠNG MỤC'!$A$2:$A$100", allow_blank=True)
    dv_cap1.error ='Dữ liệu nhập không hợp lệ'
    dv_cap1.errorTitle = 'Lỗi nhập liệu'
    dv_cap1.prompt = 'Chọn Hạng mục Cấp 1'
    dv_cap1.promptTitle = 'Hạng mục Cấp 1'
    ws_gvhv.add_data_validation(dv_cap1)
    dv_cap1.add("E2:E100")

    dv_cap2 = DataValidation(type="list", formula1="='ĐƠN GIÁ VINCONS'!$B$2:$B$9000", allow_blank=True)
    dv_cap2.error ='Dữ liệu nhập không hợp lệ'
    dv_cap2.errorTitle = 'Lỗi nhập liệu'
    dv_cap2.prompt = 'Chọn công việc/hạng mục Cấp 2'
    dv_cap2.promptTitle = 'Hạng mục Cấp 2'
    ws_gvhv.add_data_validation(dv_cap2)
    dv_cap2.add("F2:F100")

    dv_num = DataValidation(type="whole", operator="between", formula1=0, formula2=100, allow_blank=True)
    dv_num.error ='Số lượng người phải là số nguyên từ 0 đến 100'
    dv_num.errorTitle = 'Lỗi nhập số lượng'
    dv_num.prompt = 'Nhập số lượng nhân sự tham gia'
    dv_num.promptTitle = 'Số lượng nhân sự'
    ws_gvhv.add_data_validation(dv_num)
    
    dv_num.add("I2:I100")
    dv_num.add("J2:J100")
    dv_num.add("K2:K100")
    dv_num.add("L2:L100")
    dv_num.add("M2:M100")

    # Apply formulas to all formula sheets
    for sh_name in ['GIAO VIỆC HÒA VỐN', 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY', 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG', 'ĐÁNH GIÁ TỔNG DA']:
        apply_sheet_formulas(sh_name)

    # Save
    wb.save(EXCEL_PATH)
    print("Đã thêm công thức và data validation thành công!")

if __name__ == '__main__':
    modify_excel()
