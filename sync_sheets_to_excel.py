# -*- coding: utf-8 -*-
import openpyxl
import json
import urllib.request
import urllib.parse
import os
import sys
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

# Ensure UTF-8 output
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\DELL\Desktop\KPI_NhanCong_Vincons_Cleaned.xlsx"
GAS_URL = "https://script.google.com/macros/s/AKfycbw9KjUEuIKq0HYu4S8MWLvEd4cHaCNFSttPnEQOQlM0bqo_-F3VNBqfQ42CG8XH8itw/exec"

SHEETS_MAPPING = {
    'danh_sach_cong_nhan': 'DANH SÁCH CÔNG NHÂN',
    'don_gia_vincons': 'DON_GIA_VINCONS',
    'danh_muc_to': 'DANH_MUC_TO',
    'bang_luong_chuan': 'BANG_LUONG_CHUAN',
    'nhan_su_quy_luong': 'NHAN_SU_QUY_LUONG',
    'danh_gia_tong': 'DANH_GIA_TONG',
    'danh_muc_hangmuc_cv': 'DANH_MUC_HANG_MUC',
    'giao_viec_hoa_von': 'GIAO_VIEC_HOA_VON',
    'luu_tru_pgv': 'LUU_TRU_PGV',
    'bao_cao_san_luong_ngay': 'bao_cao_san_luong_ngay',
    'danh_gia_tong_trung_doi_truong': 'danh_gia_tong_trung_doi_truong',
    'danh_gia_tong_da': 'DANH_GIA_TONG_DA'
}

# Formula definitions for sheets
def apply_formulas_and_stt(ws):
    cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_indices = {name: idx for idx, name in enumerate(cols, 1) if name}
    sheet_title = ws.title
    
    for r in range(2, ws.max_row + 1):
        # 1. Update STT
        if 'STT' in col_indices:
            ws.cell(row=r, column=col_indices['STT'], value=r - 1)
            
        # 2. Update formulas
        if sheet_title == 'GIAO_VIEC_HOA_VON':
            if 'Đơn vị tính' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn vị tính'], value=f'=IF(F{r}="";"";IFERROR(VLOOKUP(F{r};\'DON_GIA_VINCONS\'!$B$2:$F$10000;3;FALSE);""))')
                
            if 'Đơn giá NC sau +30%' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn giá NC sau +30%'], value=f'=IF(F{r}="";"";IFERROR(VLOOKUP(F{r};\'DON_GIA_VINCONS\'!$B$2:$F$10000;5;FALSE);0))')
                
            if 'Tổng quỹ lương/ngày của nhóm' in col_indices:
                formula_q = (
                    f'=I{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;B{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Tổ trưởng*") + '
                    f'J{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;B{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ bậc 1") + '
                    f'K{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;B{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ bậc 2") + '
                    f'L{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;B{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ bậc 3") + '
                    f'M{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;B{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ phụ")'
                )
                ws.cell(row=r, column=col_indices['Tổng quỹ lương/ngày của nhóm'], value=formula_q)
                
            if 'Sản lượng hòa vốn 1 ngày' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng hòa vốn 1 ngày'], value=f'=IF(O{r}>0; Q{r}/O{r}; 0)')
                
            if 'Sản lượng đạt hòa vốn / 1 ngày – Cả tổ' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng đạt hòa vốn / 1 ngày – Cả tổ'], value=f'=IF(P{r}>0; H{r}/P{r}; 0)')
                
            if 'Tổng sản lượng hòa vốn cả đợt' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng sản lượng hòa vốn cả đợt'], value=f'=R{r}*P{r}')
                
            if 'Sản lượng theo khối lượng' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng theo khối lượng'], value=f'=H{r}*O{r}')
                
            if 'Sản lượng theo thực tế thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng theo thực tế thi công'], value=f'=Q{r}*P{r}')
                
            if 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch dự kiến Lãi(+)/Lỗ(-)'], value=f'=U{r}-V{r}')
                
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(W{r}=0;"";IF(W{r}>0;"An toàn - Đạt định mức";"Cảnh báo - Không đạt định mức (LỖ)"))')
                
        elif sheet_title == 'bao_cao_san_luong_ngay':
            if 'Đơn vị tính' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn vị tính'], value=f'=IF(G{r}="";"";IFERROR(VLOOKUP(G{r};\'DON_GIA_VINCONS\'!$B$2:$F$10000;3;FALSE);""))')
            if 'Đơn giá NC sau +30%' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn giá NC sau +30%'], value=f'=IF(G{r}="";"";IFERROR(VLOOKUP(G{r};\'DON_GIA_VINCONS\'!$B$2:$F$10000;5;FALSE);0))')
            if 'Số ngày thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Số ngày thi công'], value=f'=IF(OR(D{r}="";E{r}="");"";E{r}-D{r}+1)')
            if 'Tổng quỹ lương/ngày của nhóm' in col_indices:
                formula_q = (
                    f'=J{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;C{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Tổ trưởng*") + '
                    f'K{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;C{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ bậc 1") + '
                    f'L{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;C{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ bậc 2") + '
                    f'M{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;C{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ bậc 3") + '
                    f'N{r}*SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;C{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"Thợ phụ")'
                )
                ws.cell(row=r, column=col_indices['Tổng quỹ lương/ngày của nhóm'], value=formula_q)
            if 'Sản lượng theo thực tế thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng theo thực tế thi công'], value=f'=R{r}*Q{r}')
            if 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch dự kiến Lãi(+)/Lỗ(-)'], value=f'=(I{r}*P{r})-S{r}')
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(T{r}=0;"";IF(T{r}>0;"An toàn - Đạt định mức";"Cảnh báo - Không đạt định mức (LỖ)"))')
                
        elif sheet_title == 'danh_gia_tong_trung_doi_truong':
            if 'Số ngày thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Số ngày thi công'], value=f'=IF(OR(D{r}="";E{r}="");"";E{r}-D{r}+1)')
            if 'Tổng quỹ lương chi trả đến hiện tại của tổ' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng quỹ lương chi trả đến hiện tại của tổ'], value=f'=SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$B$2:$B$5000;C{r};\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"<>Tổ trưởng*")*F{r}')
            if 'Tổng sản lượng theo khối lượng' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng sản lượng theo khối lượng'], value=f'=SUMPRODUCT((\'bao_cao_san_luong_ngay\'!$C$2:$C$5000=C{r})*(\'bao_cao_san_luong_ngay\'!$I$2:$I$5000)*(\'bao_cao_san_luong_ngay\'!$P$2:$P$5000))')
            if 'Chênh lệch Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch Lãi(+)/Lỗ(-)'], value=f'=H{r}-G{r}')
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(I{r}=0;"";IF(I{r}>0;"Lãi";"Lỗ"))')
                
        elif sheet_title == 'DANH_GIA_TONG_DA':
            if 'Số ngày thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Số ngày thi công'], value=f'=IF(OR(C{r}="";D{r}="");"";D{r}-C{r}+1)')
            if 'Tổng quỹ lương chi trả đến hiện tại' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng quỹ lương chi trả đến hiện tại'], value=f'=SUMIFS(\'NHAN_SU_QUY_LUONG\'!$H$2:$H$5000;\'NHAN_SU_QUY_LUONG\'!$E$2:$E$5000;"<>Tổ trưởng*")*E{r}')
            if 'Tổng sản lượng theo khối lượng' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng sản lượng theo khối lượng'], value=f'=SUMPRODUCT(\'bao_cao_san_luong_ngay\'!$I$2:$I$5000*\'bao_cao_san_luong_ngay\'!$P$2:$P$5000)')
            if 'Chênh lệch Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch Lãi(+)/Lỗ(-)'], value=f'=G{r}-F{r}')
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(H{r}=0;"";IF(H{r}>0;"Lãi";"Lỗ"))')

def download_sheet_data(sheet_name):
    print(f"Đang tải dữ liệu từ sheet: '{sheet_name}'...")
    try:
        req_url = GAS_URL + '?action=getSheet&name=' + urllib.parse.quote(sheet_name)
        req = urllib.request.Request(req_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req) as res:
            data = json.loads(res.read().decode('utf-8'))
            return data
    except Exception as e:
        print(f"Lỗi tải dữ liệu cho '{sheet_name}': {e}")
        return None

def main():
    # 1. Check if backup directory is needed
    dir_name = os.path.dirname(EXCEL_PATH)
    if not os.path.exists(dir_name):
        os.makedirs(dir_name)

    # Backup existing file if any
    if os.path.exists(EXCEL_PATH):
        backup_path = EXCEL_PATH + f".backup_{int(os.path.getmtime(EXCEL_PATH))}"
        try:
            import shutil
            shutil.copyfile(EXCEL_PATH, backup_path)
            print(f"Đã sao lưu tệp Excel hiện tại vào {backup_path}")
        except Exception as e:
            print(f"Cảnh báo: Không thể tạo bản sao lưu: {e}")

    # Create a new workbook
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 2. Iterate through all worksheets and write to the Excel file
    for sheet_id, sheet_name in SHEETS_MAPPING.items():
        records = download_sheet_data(sheet_name)
        if records is None:
            print(f"Bỏ qua sheet {sheet_name} do lỗi tải.")
            continue
            
        if isinstance(records, dict) and records.get('isGrid'):
            print(f"Đã tải {len(records['values'])} dòng dạng lưới.")
            ws = wb.create_sheet(title=sheet_name)
            for r_idx, row in enumerate(records['values'], 1):
                for c_idx, val in enumerate(row, 1):
                    if val is None or val == "":
                        ws.cell(row=r_idx, column=c_idx, value=None)
                    else:
                        try:
                            if '.' in str(val):
                                ws.cell(row=r_idx, column=c_idx, value=float(val))
                            else:
                                ws.cell(row=r_idx, column=c_idx, value=int(val))
                        except ValueError:
                            ws.cell(row=r_idx, column=c_idx, value=str(val))
        else:
            print(f"Đã tải {len(records)} dòng.")
            ws = wb.create_sheet(title=sheet_name)
            if len(records) > 0:
                df = pd.DataFrame(records)
                
                # Write headers
                headers = list(df.columns)
                ws.append(headers)
                
                # Write data rows
                for _, row in df.iterrows():
                    row_vals = []
                    for val in row:
                        if val is None or val == "":
                            row_vals.append(None)
                        else:
                            try:
                                if '.' in str(val):
                                    row_vals.append(float(val))
                                else:
                                    row_vals.append(int(val))
                            except ValueError:
                                row_vals.append(str(val))
                    ws.append(row_vals)
            else:
                # Empty sheet - write default headers if we have a fallback
                schemas = {
                    'DANH SÁCH CÔNG NHÂN': ['STT', 'Họ và tên', 'Số điện thoại', 'Tổ thi công', 'Chức danh', 'Ghi chú'],
                    'DON_GIA_VINCONS': ['STT', 'Hạng mục thi công (Nội dung công việc)', 'Đơn vị tính', 'Đơn giá khoán tổ NC', 'Đơn giá NC sau +30%', 'Ghi chú'],
                    'DANH_MUC_TO': ['STT', 'Tên Tổ', 'Tổ trưởng', 'Số điện thoại', 'Ghi chú'],
                    'BANG_LUONG_CHUAN': ['STT', 'Chức danh', 'Mức lương chuẩn / Ngày', 'Ghi chú'],
                    'NHAN_SU_QUY_LUONG': ['STT', 'Họ và tên', 'Chức danh', 'Tổ thi công', 'Mức lương chuẩn / Ngày', 'Hệ số lương', 'Lương 1 ngày thực tế', 'Ghi chú'],
                    'DANH_GIA_TONG': ['STT', 'Tổ thi công', 'Ngày bắt đầu', 'Ngày kết thúc', 'Số ngày thi công', 'Tổng lương', 'Tổng sản lượng theo khối lượng', 'Chênh lệch Lãi(+)/Lỗ(-)', 'ĐÁNH GIÁ & CẢNH BÁO'],
                    'DANH_MUC_HANG_MUC': ['STT', 'Hạng mục thi công', 'Ghi chú'],
                    'GIAO_VIEC_HOA_VON': ['STT', 'Tổ thi công', 'Ngày bắt đầu', 'Ngày kết thúc', 'Hạng mục thi công (Cấp 1)', 'Hạng mục thi công (Cấp 2)', 'Vị trí thi công', 'Tổng khối lượng yêu cầu', 'SL Tổ trưởng tham gia', 'SL Thợ bậc 1 tham gia', 'SL Thợ bậc 2 tham gia', 'SL Thợ bậc 3 tham gia', 'SL Thợ phụ tham gia', 'Đơn vị tính', 'Đơn giá NC sau +30%', 'Số ngày thi công', 'Tổng quỹ lương/ngày của nhóm', 'Sản lượng hòa vốn 1 ngày', 'Sản lượng đạt hòa vốn / 1 ngày – Cả tổ', 'Tổng sản lượng hòa vốn cả đợt', 'Sản lượng theo khối lượng', 'Sản lượng theo thực tế thi công', 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)', 'ĐÁNH GIÁ & CẢNH BÁO'],
                    'LUU_TRU_PGV': ['STT', 'Mã PGV', 'Tổ thi công', 'Ngày bắt đầu', 'Ngày kết thúc', 'Trạng thái', 'Ghi chú'],
                    'bao_cao_san_luong_ngay': ['Trung đội trưởng QL', 'STT', 'Tổ thi công', 'Ngày bắt đầu', 'Ngày kết thúc', 'Hạng mục thi công (Cấp 1)', 'Hạng mục thi công (Cấp 2)', 'Vị trí thi công', 'Tổng khối lượng thực hiện', 'SL Tổ trưởng tham gia', 'SL Thợ bậc 1 tham gia', 'SL Thợ bậc 2 tham gia', 'SL Thợ bậc 3 tham gia', 'SL Thợ phụ tham gia', 'Đơn vị tính', 'Đơn giá NC sau +30%', 'Số ngày thi công', 'Tổng quỹ lương/ngày của nhóm', 'Sản lượng theo thực tế thi công', 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)', 'ĐÁNH GIÁ & CẢNH BÁO'],
                    'danh_gia_tong_trung_doi_truong': ['STT', 'Trung đội trưởng QL', 'Tổ thi công', 'Ngày bắt đầu', 'Ngày kết thúc', 'Số ngày thi công', 'Tổng quỹ lương chi trả đến hiện tại của tổ', 'Tổng sản lượng theo khối lượng', 'Chênh lệch Lãi(+)/Lỗ(-)', 'ĐÁNH GIÁ & CẢNH BÁO'],
                    'DANH_GIA_TONG_DA': ['STT', 'Dự án', 'Ngày bắt đầu', 'Ngày kết thúc', 'Số ngày thi công', 'Tổng quỹ lương chi trả đến hiện tại', 'Tổng sản lượng theo khối lượng', 'Chênh lệch Lãi(+)/Lỗ(-)', 'ĐÁNH GIÁ & CẢNH BÁO']
                }
                ws.append(schemas.get(sheet_name, []))

        # Apply formulas
        if sheet_name in ['GIAO_VIEC_HOA_VON', 'bao_cao_san_luong_ngay', 'danh_gia_tong_trung_doi_truong', 'DANH_GIA_TONG_DA']:
            apply_formulas_and_stt(ws)
            print(f"Đã áp dụng công thức tự động cho sheet: '{sheet_name}'")

    # 3. Save Excel Workbook
    try:
        wb.save(EXCEL_PATH)
        print(f"\n[OK] Đồng bộ thành công tất cả các sheet trực tuyến vào: {EXCEL_PATH}!")
    except PermissionError:
        print(f"\n[LỖI] Không thể ghi đè file Excel. Vui lòng đóng tệp Excel trên máy tính của bạn trước!")
    except Exception as e:
        print(f"\n[LỖI] Lỗi lưu file: {e}")

if __name__ == '__main__':
    main()
