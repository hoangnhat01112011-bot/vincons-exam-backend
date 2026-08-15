import openpyxl
import json
import urllib.request
import os
import sys
import datetime

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = r"C:\Users\DELL\Desktop\KPI_NhanCong_Vincons_Cleaned.xlsx"
GAS_URL = "https://script.google.com/macros/s/AKfycbx1ALjAaojX7tDhxVrM26CQeu4FL4RQ4K7W7EoABTM/dev"

SHEETS_MAPPING = {
    'danh_sach_cong_nhan': 'DANH SÁCH CÔNG NHÂN',
    'don_gia_vincons': 'DON_GIA_VINCONS',
    'danh_muc_to': 'DANH_MUC_TO',
    'bang_luong_chuan': 'BANG_LUONG_CHUAN',
    'nhan_su_quy_luong': 'NHAN_SU_QUY_LUONG',
    'danh_gia_tong': 'DANH_GIA_TONG',
    'danh_muc_hangmuc_cv': 'DANH_MUC_HANG_MUC',
    'giao_viec_hoa_von': 'GIAO_VIEC_HOA_VON',
    'luu_tru_pgv': 'LUU_TRU_PGV',
    'bao_cao_san_luong_ngay': 'bao_cao_san_luong_ngay',
    'danh_gia_tong_trung_doi_truong': 'danh_gia_tong_trung_doi_truong',
    'danh_gia_tong_da': 'DANH_GIA_TONG_DA'
}

def upload_sheet(sheet_id, sheet_name):
    print(f"Đang đọc sheet: '{sheet_name}'...")
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
        if sheet_name not in wb.sheetnames:
            print(f"Lỗi: Không tìm thấy sheet '{sheet_name}' trong file Excel.")
            return
        ws = wb[sheet_name]
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            print(f"Sheet '{sheet_name}' rỗng.")
            return
            
        if sheet_id == 'luu_tru_pgv':
            # Send raw 2D grid
            grid_vals = []
            for row in rows:
                row_vals = []
                for val in row:
                    if val is None:
                        row_vals.append("")
                    elif isinstance(val, datetime.datetime):
                        row_vals.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                    else:
                        row_vals.append(val)
                grid_vals.append(row_vals)
            payload = {
                "action": "importData",
                "sheet": sheet_id,
                "isGrid": True,
                "data": grid_vals
            }
            print(f"Đang tải lưới {len(grid_vals)}x{len(grid_vals[0]) if grid_vals else 0} lên Google Sheets...")
        else:
            headers = rows[0]
            records = []
            for row in rows[1:]:
                if all(v is None or v == "" for v in row):
                    continue
                obj = {}
                for h, val in zip(headers, row):
                    if h is None:
                        continue
                    if val is None:
                        val = ""
                    elif isinstance(val, datetime.datetime):
                        val = val.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(val, (int, float)):
                        pass
                    else:
                        val = str(val)
                    obj[h] = val
                records.append(obj)
                
            payload = {
                "action": "importData",
                "sheet": sheet_id,
                "isGrid": False,
                "data": records
            }
            print(f"Đang tải {len(records)} bản ghi lên Google Sheets...")
    
    try:
        data = json.dumps(payload, ensure_ascii=False).encode('utf-8')
        req = urllib.request.Request(
            GAS_URL, 
            data=data,
            headers={'Content-Type': 'application/json', 'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req) as res:
            res_data = json.loads(res.read().decode('utf-8'))
            print(f"Kết quả: {res_data.get('message', 'Không có thông báo')}\n")
    except Exception as e:
        print(f"Lỗi gọi API: {e}\n")

if __name__ == '__main__':
    if not os.path.exists(EXCEL_PATH):
        print(f"Lỗi: Không tìm thấy file Excel tại {EXCEL_PATH}")
    else:
        for sheet_id, sheet_name in SHEETS_MAPPING.items():
            upload_sheet(sheet_id, sheet_name)
        print("Đồng bộ hoàn tất!")
