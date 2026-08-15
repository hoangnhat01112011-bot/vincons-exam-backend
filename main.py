import http.server
import socketserver
import webbrowser
import threading
import os
import sys
import json
import csv
import hashlib
from urllib.parse import parse_qs, urlparse
import pandas as pd
import openpyxl
PORT = int(os.environ.get('PORT', 8000))
DIRECTORY = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(DIRECTORY, 'data')

SHEETS_MAPPING = {
    'danh_sach_cong_nhan': 'DANH SÁCH CÔNG NHÂN',
    'don_gia_vincons': 'ĐƠN GIÁ VINCONS',
    'danh_muc_to': 'DANH MỤC TỔ',
    'bang_luong_chuan': 'BẢNG LƯƠNG CHUẨN',
    'nhan_su_quy_luong': 'NHÂN SỰ QUỸ LƯƠNG',
    'danh_gia_tong': 'ĐÁNH GIÁ TỔNG',
    'danh_muc_hangmuc_cv': 'DANH MỤC HẠNG MỤC',
    'giao_viec_hoa_von': 'GIAO VIỆC HÒA VỐN',
    'luu_tru_pgv': 'LƯU TRỮ PGV',
    'bao_cao_san_luong_ngay': 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY',
    'danh_gia_tong_trung_doi_truong': 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG',
    'danh_gia_tong_da': 'ĐÁNH GIÁ TỔNG DA'
}

def get_cleaned_excel_path():
    paths_to_check = [
        os.path.join(DIRECTORY, "KPI_NhanCong_Vincons_Cleaned.xlsx"),
        os.path.join(DATA_DIR, "KPI_NhanCong_Vincons_Cleaned.xlsx"),
        os.path.join(os.path.expanduser("~"), "Desktop", "KPI_NhanCong_Vincons_Cleaned.xlsx"),
        os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop", "KPI_NhanCong_Vincons_Cleaned.xlsx"),
        r"C:\Users\DELL\Desktop\KPI_NhanCong_Vincons_Cleaned.xlsx"
    ]
    for path in paths_to_check:
        if os.path.exists(path):
            return path
    return os.path.join(os.path.expanduser("~"), "Desktop", "KPI_NhanCong_Vincons_Cleaned.xlsx")

CLEANED_EXCEL = get_cleaned_excel_path()
ADMIN_ONLY_SHEETS = [
    'don_gia_vincons',
    'bang_luong_chuan',
    'nhan_su_quy_luong',
    'danh_gia_tong'
]

def apply_formulas_and_stt(ws):
    cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_indices = {name: idx for idx, name in enumerate(cols, 1) if name}
    sheet_title = ws.title
    
    for r in range(2, ws.max_row + 1):
        # 1. Update STT
        if 'STT' in col_indices:
            ws.cell(row=r, column=col_indices['STT'], value=r - 1)
            
        # 2. Update formulas
        if sheet_title == 'GIAO VIỆC HÒA VỐN':
            if 'Đơn vị tính' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn vị tính'], value=f'=IF(F{r}="";"";IFERROR(VLOOKUP(F{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;3;FALSE);""))')
                
            if 'Đơn giá NC sau +30%' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn giá NC sau +30%'], value=f'=IF(F{r}="";"";IFERROR(VLOOKUP(F{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;5;FALSE);0))')
                
            if 'Tổng quỹ lương/ngày của nhóm' in col_indices:
                formula_q = (
                    f'=I{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Tổ trưởng*") + '
                    f'J{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 1") + '
                    f'K{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 2") + '
                    f'L{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 3") + '
                    f'M{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;B{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ phụ")'
                )
                ws.cell(row=r, column=col_indices['Tổng quỹ lương/ngày của nhóm'], value=formula_q)
                
            if 'Sản lượng hòa vốn 1 ngày' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng hòa vốn 1 ngày'], value=f'=IF(O{r}>0; Q{r}/O{r}; 0)')
                
            if 'Sản lượng đạt hòa vốn / 1 ngày – Cả tổ' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng đạt hòa vốn / 1 ngày – Cả tổ'], value=f'=IF(P{r}>0; H{r}/P{r}; 0)')
                
            if 'Tổng sản lượng hòa vốn cả đợt' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng sản lượng hòa vốn cả đợt'], value=f'=R{r}*P{r}')
                
            if 'Sản lượng theo khối lượng' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng theo khối lượng'], value=f'=H{r}*O{r}')
                
            if 'Sản lượng theo thực tế thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng theo thực tế thi công'], value=f'=Q{r}*P{r}')
                
            if 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch dự kiến Lãi(+)/Lỗ(-)'], value=f'=U{r}-V{r}')
                
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(W{r}=0;"";IF(W{r}>0;"An toàn - Đạt định mức";"Cảnh báo - Không đạt định mức (LỖ)"))')
                
        elif sheet_title == 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY':
            if 'Đơn vị tính' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn vị tính'], value=f'=IF(G{r}="";"";IFERROR(VLOOKUP(G{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;3;FALSE);""))')
            if 'Đơn giá NC sau +30%' in col_indices:
                ws.cell(row=r, column=col_indices['Đơn giá NC sau +30%'], value=f'=IF(G{r}="";"";IFERROR(VLOOKUP(G{r};\'ĐƠN GIÁ VINCONS\'!$B$2:$F$10000;5;FALSE);0))')
            if 'Số ngày thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Số ngày thi công'], value=f'=IF(OR(D{r}="";E{r}="");"";E{r}-D{r}+1)')
            if 'Tổng quỹ lương/ngày của nhóm' in col_indices:
                formula_q = (
                    f'=J{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Tổ trưởng*") + '
                    f'K{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 1") + '
                    f'L{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 2") + '
                    f'M{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ bậc 3") + '
                    f'N{r}*SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"Thợ phụ")'
                )
                ws.cell(row=r, column=col_indices['Tổng quỹ lương/ngày của nhóm'], value=formula_q)
            if 'Sản lượng theo thực tế thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Sản lượng theo thực tế thi công'], value=f'=R{r}*Q{r}')
            if 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch dự kiến Lãi(+)/Lỗ(-)'], value=f'=(I{r}*P{r})-S{r}')
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(T{r}=0;"";IF(T{r}>0;"An toàn - Đạt định mức";"Cảnh báo - Không đạt định mức (LỖ)"))')
                
        elif sheet_title == 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG':
            if 'Số ngày thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Số ngày thi công'], value=f'=IF(OR(D{r}="";E{r}="");"";E{r}-D{r}+1)')
            if 'Tổng quỹ lương chi trả đến hiện tại của tổ' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng quỹ lương chi trả đến hiện tại của tổ'], value=f'=SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$B$2:$B$5000;C{r};\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"<>Tổ trưởng*")*F{r}')
            if 'Tổng sản lượng theo khối lượng' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng sản lượng theo khối lượng'], value=f'=SUMPRODUCT((\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$C$2:$C$5000=C{r})*(\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$I$2:$I$5000)*(\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$P$2:$P$5000))')
            if 'Chênh lệch Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch Lãi(+)/Lỗ(-)'], value=f'=H{r}-G{r}')
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(I{r}=0;"";IF(I{r}>0;"Lãi";"Lỗ"))')
                
        elif sheet_title == 'ĐÁNH GIÁ TỔNG DA':
            if 'Số ngày thi công' in col_indices:
                ws.cell(row=r, column=col_indices['Số ngày thi công'], value=f'=IF(OR(C{r}="";D{r}="");"";D{r}-C{r}+1)')
            if 'Tổng quỹ lương chi trả đến hiện tại' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng quỹ lương chi trả đến hiện tại'], value=f'=SUMIFS(\'NHÂN SỰ QUỸ LƯƠNG\'!$H$2:$H$5000;\'NHÂN SỰ QUỸ LƯƠNG\'!$E$2:$E$5000;"<>Tổ trưởng*")*E{r}')
            if 'Tổng sản lượng theo khối lượng' in col_indices:
                ws.cell(row=r, column=col_indices['Tổng sản lượng theo khối lượng'], value=f'=SUMPRODUCT(\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$I$2:$I$5000*\'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY\'!$P$2:$P$5000)')
            if 'Chênh lệch Lãi(+)/Lỗ(-)' in col_indices:
                ws.cell(row=r, column=col_indices['Chênh lệch Lãi(+)/Lỗ(-)'], value=f'=G{r}-F{r}')
            if 'ĐÁNH GIÁ & CẢNH BÁO' in col_indices:
                ws.cell(row=r, column=col_indices['ĐÁNH GIÁ & CẢNH BÁO'], value=f'=IF(H{r}=0;"";IF(H{r}>0;"Lãi";"Lỗ"))')

GAS_URL = "https://script.google.com/macros/s/AKfycbwwXK6vhnJe4EDqiYjvP0WpUDfjs98LJ1FWbRnom1Z7aa_-VQonYpIj1WTgT0SGjDs1/exec"

def call_gas_api(params=None, post_data=None):
    import urllib.request
    import urllib.parse
    
    sheet_maps = {
        'danh_sach_cong_nhan': 'DANH SÁCH CÔNG NHÂN',
        'don_gia_vincons': 'ĐƠN GIÁ VINCONS',
        'danh_muc_to': 'DANH MỤC TỔ',
        'bang_luong_chuan': 'BẢNG LƯƠNG CHUẨN',
        'nhan_su_quy_luong': 'NHÂN SỰ QUỸ LƯƠNG',
        'danh_gia_tong': 'ĐÁNH GIÁ TỔNG',
        'danh_muc_hangmuc_cv': 'DANH MỤC HẠNG MỤC',
        'giao_viec_hoa_von': 'GIAO VIỆC HÒA VỐN',
        'luu_tru_pgv': 'LƯU TRỮ PGV',
        'bao_cao_san_luong_ngay': 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY',
        'danh_gia_tong_trung_doi_truong': 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG',
        'danh_gia_tong_da': 'ĐÁNH GIÁ TỔNG DA'
    }
    
    if params and 'name' in params:
        params['name'] = sheet_maps.get(params['name'], params['name'])
        
    if post_data and 'sheet' in post_data:
        post_data['sheet'] = sheet_maps.get(post_data['sheet'], post_data['sheet'])

    url = GAS_URL
    if params:
        query_string = urllib.parse.urlencode(params)
        url += ("&" if "?" in url else "?") + query_string
        
    req_headers = {'User-Agent': 'Mozilla/5.0'}
    
    if post_data is not None:
        data_bytes = json.dumps(post_data, ensure_ascii=False).encode('utf-8')
        req_headers['Content-Type'] = 'application/json'
        req = urllib.request.Request(url, data=data_bytes, headers=req_headers, method='POST')
    else:
        req = urllib.request.Request(url, headers=req_headers, method='GET')
        
    with urllib.request.urlopen(req) as res:
        return json.loads(res.read().decode('utf-8'))

# Ensure UTF-8 output encoding for Windows console
if hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except Exception:
        pass

RESULTS_JSON = os.path.join(DATA_DIR, 'results.json')
RESULTS_CSV = os.path.join(DATA_DIR, 'results.csv')
USERS_JSON = os.path.join(DATA_DIR, 'users.json')

def ensure_data_dir():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

def load_users():
    ensure_data_dir()
    if not os.path.exists(USERS_JSON):
        # Create empty list if it doesn't exist
        try:
            with open(USERS_JSON, 'w', encoding='utf-8') as f:
                json.dump([], f)
        except Exception:
            pass
        return []
    try:
        with open(USERS_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading users: {e}")
        return []

def save_users(users):
    ensure_data_dir()
    try:
        with open(USERS_JSON, 'w', encoding='utf-8') as f:
            json.dump(users, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving users: {e}")

def load_results():
    ensure_data_dir()
    if not os.path.exists(RESULTS_JSON):
        return []
    try:
        with open(RESULTS_JSON, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading results: {e}")
        return []

def save_results_to_files(results):
    ensure_data_dir()
    # Save to JSON
    with open(RESULTS_JSON, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    # Save to CSV with UTF-8 BOM for Microsoft Excel compatibility
    try:
        with open(RESULTS_CSV, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                "Mã bản ghi",
                "Họ và tên",
                "SBD/SĐT",
                "Vị trí ứng tuyển",
                "Phần thi",
                "Điểm số",
                "Số câu đúng",
                "Tổng số câu",
                "Tỷ lệ (%)",
                "Kết quả",
                "Thời gian nộp"
            ])
            for item in results:
                cand = item.get('candidate', {})
                score_str = f"{item.get('correctCount', 0)}/{item.get('totalCount', 0)}"
                pct = item.get('percentage', 0)
                writer.writerow([
                    item.get('id', ''),
                    cand.get('name', ''),
                    cand.get('id', ''),
                    cand.get('job', ''),
                    cand.get('examType', ''),
                    score_str,
                    item.get('correctCount', 0),
                    item.get('totalCount', 0),
                    f"{pct:.1f}%",
                    "ĐẠT" if item.get('isPass') else "CHƯA ĐẠT",
                    item.get('submitTime', '')
                ])
    except Exception as e:
        print(f"Error saving CSV: {e}")

class Handler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DIRECTORY, **kwargs)

    def end_headers(self):
        self.send_header('Cache-Control', 'no-store, no-cache, must-revalidate, max-age=0')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()

    def send_json(self, data, code=200):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(code)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, DELETE, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_GET(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        
        if parsed.path == '/api/kpi/sheet':
            sheet_name = query.get('name', [None])[0]
            if not sheet_name:
                self.send_json({"status": "error", "message": "Missing sheet name"}, code=400)
                return
            excel_path = get_cleaned_excel_path()
            if not os.path.exists(excel_path):
                try:
                    gas_data = call_gas_api(params={"action": "getSheet", "name": sheet_name})
                    self.send_json(gas_data)
                except Exception as e:
                    self.send_json({"status": "error", "message": f"Không tìm thấy tệp Excel cục bộ và lỗi kết nối Google Sheets: {str(e)}"}, code=500)
                return
            mapped_sheet = SHEETS_MAPPING.get(sheet_name)
            if not mapped_sheet:
                self.send_json({"status": "error", "message": f"Sheet '{sheet_name}' không hợp lệ!"}, code=400)
                return
            try:
                if sheet_name == 'luu_tru_pgv':
                    wb = openpyxl.load_workbook(excel_path, data_only=True)
                    ws = wb[mapped_sheet]
                    values = []
                    for row in ws.iter_rows(values_only=True):
                        row_vals = []
                        for val in row:
                            if val is None:
                                row_vals.append("")
                            elif hasattr(val, 'strftime'):
                                row_vals.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                            else:
                                row_vals.append(val)
                        values.append(row_vals)
                    self.send_json({
                        "isGrid": True,
                        "values": values
                    })
                else:
                    df = pd.read_excel(excel_path, sheet_name=mapped_sheet)
                    for col in df.columns:
                        if pd.api.types.is_datetime64_any_dtype(df[col]):
                            df[col] = df[col].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
                    df = df.fillna('')
                    records = df.to_dict(orient='records')
                    self.send_json(records)
            except Exception as e:
                self.send_json({"status": "error", "message": f"Lỗi đọc sheet {mapped_sheet}: {str(e)}"}, code=500)
            return

        elif parsed.path == '/api/results':
            results = load_results()
            user_id = query.get('userId', [None])[0]
            role = query.get('role', [None])[0]
            
            if role == 'admin':
                self.send_json(results)
            elif role == 'candidate' and user_id:
                # Filter results where candidate.id matches the candidate's logged in username (phone/email)
                filtered = [r for r in results if str(r.get('candidate', {}).get('id', '')).strip().lower() == user_id.strip().lower()]
                self.send_json(filtered)
            else:
                self.send_json([])
            return
            
        elif parsed.path == '/api/list-pdfs':
            ensure_data_dir()
            pdf_files = []
            if os.path.exists(DATA_DIR):
                for fname in os.listdir(DATA_DIR):
                    if fname.lower().endswith('.pdf'):
                        pdf_files.append({
                            "name": fname,
                            "url": f"/data/{fname}"
                        })
            self.send_json(pdf_files)
            return
            
        elif parsed.path == '/api/export-csv':
            ensure_data_dir()
            user_id = query.get('userId', [None])[0]
            role = query.get('role', [None])[0]
            
            if not role or (role != 'admin' and role != 'candidate'):
                self.send_json({"status": "error", "message": "Unauthorized"}, code=403)
                return
                
            results = load_results()
            if role == 'candidate' and user_id:
                results = [r for r in results if str(r.get('candidate', {}).get('id', '')).strip().lower() == user_id.strip().lower()]
            
            # Generate CSV bytes on the fly
            try:
                import io
                f = io.StringIO()
                f.write('\ufeff') # UTF-8 BOM
                writer = csv.writer(f)
                writer.writerow([
                    "Mã bản ghi",
                    "Họ và tên",
                    "SBD/SĐT",
                    "Vị trí ứng tuyển",
                    "Phần thi",
                    "Điểm số",
                    "Số câu đúng",
                    "Tổng số câu",
                    "Tỷ lệ (%)",
                    "Kết quả",
                    "Thời gian nộp"
                ])
                for item in results:
                    cand = item.get('candidate', {})
                    score_str = f"{item.get('correctCount', 0)}/{item.get('totalCount', 0)}"
                    pct = item.get('percentage', 0)
                    writer.writerow([
                        item.get('id', ''),
                        cand.get('name', ''),
                        cand.get('id', ''),
                        cand.get('job', ''),
                        cand.get('examType', ''),
                        score_str,
                        item.get('correctCount', 0),
                        item.get('totalCount', 0),
                        f"{pct:.1f}%",
                        "ĐẠT" if item.get('isPass') else "CHƯA ĐẠT",
                        item.get('submitTime', '')
                    ])
                content = f.getvalue().encode('utf-8-sig')
                
                self.send_response(200)
                self.send_header('Content-Type', 'text/csv; charset=utf-8')
                self.send_header('Content-Disposition', 'attachment; filename="vincons_exam_results.csv"')
                self.send_header('Content-Length', str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=500)
                return
        
        # Standard static file handling
        return super().do_GET()

    def do_POST(self):
        parsed = urlparse(self.path)
        
        if parsed.path == '/api/auth/register':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                username = data.get('username', '').strip()
                password = data.get('password', '')
                name = data.get('name', '').strip()
                role = data.get('role', 'candidate').strip()
                admin_code = data.get('adminCode', '').strip()
                
                if not username or not password or not name:
                    self.send_json({"status": "error", "message": "Vui lòng nhập đầy đủ thông tin bắt buộc!"}, code=400)
                    return
                
                if role == 'admin' and admin_code != 'VINCONS2026':
                    self.send_json({"status": "error", "message": "Mã kích hoạt giám thị không chính xác!"}, code=400)
                    return
                
                users = load_users()
                # Check if username already exists
                if any(u.get('username', '').lower() == username.lower() for u in users):
                    self.send_json({"status": "error", "message": "Số điện thoại hoặc Gmail này đã được đăng ký!"}, code=400)
                    return
                
                # Hash password
                hashed_pw = hashlib.sha256(password.encode('utf-8')).hexdigest()
                new_user = {
                    "username": username,
                    "password": hashed_pw,
                    "name": name,
                    "role": role
                }
                users.append(new_user)
                save_users(users)
                self.send_json({"status": "success", "message": "Đăng ký tài khoản thành công!"})
            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=400)
            return

        elif parsed.path == '/api/auth/login':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                username = data.get('username', '').strip()
                password = data.get('password', '')
                
                if not username or not password:
                    self.send_json({"status": "error", "message": "Vui lòng nhập tài khoản và mật khẩu!"}, code=400)
                    return
                
                hashed_pw = hashlib.sha256(password.encode('utf-8')).hexdigest()
                users = load_users()
                
                found_user = None
                for u in users:
                    if u.get('username', '').lower() == username.lower() and u.get('password', '') == hashed_pw:
                        found_user = u
                        break
                
                if found_user:
                    self.send_json({
                        "status": "success",
                        "message": "Đăng nhập thành công!",
                        "user": {
                            "username": found_user["username"],
                            "name": found_user["name"],
                            "role": found_user["role"]
                        }
                    })
                else:
                    self.send_json({"status": "error", "message": "Số điện thoại/Gmail hoặc mật khẩu không đúng!"}, code=400)
            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=400)
            return

        elif parsed.path == '/api/save-result':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                results = load_results()
                
                # Check if item with same ID already exists, update or insert
                item_id = data.get('id')
                if not item_id:
                    item_id = f"res_{int(threading.get_ident())}_{len(results) + 1}"
                    data['id'] = item_id

                # Upsert into list
                existing_idx = next((i for i, r in enumerate(results) if r.get('id') == item_id), -1)
                if existing_idx >= 0:
                    results[existing_idx] = data
                else:
                    results.insert(0, data) # Newest first

                save_results_to_files(results)
                self.send_json({"status": "success", "message": "Đã lưu kết quả bài thi thành công", "id": item_id})
            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=400)
            return

        elif parsed.path == '/api/kpi/save-record':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                sheet_name = data.get('sheet')
                row_index = int(data.get('rowIndex', -1))
                record = data.get('record')
                role = data.get('role', 'candidate')

                if not sheet_name or not record:
                    self.send_json({"status": "error", "message": "Thiếu dữ liệu sheet hoặc bản ghi!"}, code=400)
                    return

                if sheet_name in ADMIN_ONLY_SHEETS and role != 'admin':
                    self.send_json({"status": "error", "message": "Bạn không có quyền chỉnh sửa mục này! Cần quyền Admin."}, code=403)
                    return

                excel_path = get_cleaned_excel_path()
                
                # Write to Google Sheets first
                gas_res = None
                try:
                    post_body = {
                        "action": "saveRecord",
                        "sheet": sheet_name,
                        "rowIndex": row_index,
                        "record": record
                    }
                    gas_res = call_gas_api(post_data=post_body)
                    if gas_res.get('status') != 'success':
                        self.send_json(gas_res, code=400)
                        return
                except Exception as e:
                    self.send_json({"status": "error", "message": f"Lỗi đồng bộ Google Sheets: {str(e)}"}, code=500)
                    return
                    
                if not os.path.exists(excel_path):
                    self.send_json(gas_res)
                    return

                mapped_sheet = SHEETS_MAPPING.get(sheet_name)
                if not mapped_sheet:
                    self.send_json({"status": "error", "message": f"Sheet '{sheet_name}' không hợp lệ!"}, code=400)
                    return

                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb[mapped_sheet]
                    
                    # Columns in the sheet
                    cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                    col_indices = {name: idx for idx, name in enumerate(cols, 1) if name}
                    
                    if row_index >= 0:
                        r = row_index + 2
                    else:
                        r = ws.max_row + 1
                        
                    for col_name, col_idx in col_indices.items():
                        # Skip formula cols in GIAO VIỆC HÒA VỐN and others (they are handled in apply_formulas_and_stt)
                        if mapped_sheet == 'GIAO VIỆC HÒA VỐN':
                            formula_cols = [
                                'STT', 'Đơn vị tính', 'Đơn giá NC sau +30%', 
                                'Tổng quỹ lương/ngày của nhóm', 'Sản lượng hòa vốn 1 ngày', 
                                'Sản lượng đạt hòa vốn / 1 ngày – Cả tổ', 'Tổng sản lượng hòa vốn cả đợt', 
                                'Sản lượng theo khối lượng', 'Sản lượng theo thực tế thi công', 
                                'Chênh lệch dự kiến Lãi(+)/Lỗ(-)', 'ĐÁNH GIÁ & CẢNH BÁO'
                            ]
                            if col_name in formula_cols:
                                continue
                        elif mapped_sheet == 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY':
                            formula_cols = [
                                'STT', 'Đơn vị tính', 'Đơn giá NC sau +30%', 
                                'Số ngày thi công', 'Tổng quỹ lương/ngày của nhóm', 
                                'Sản lượng theo thực tế thi công', 'Chênh lệch dự kiến Lãi(+)/Lỗ(-)', 
                                'ĐÁNH GIÁ & CẢNH BÁO'
                            ]
                            if col_name in formula_cols:
                                continue
                        elif mapped_sheet == 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG':
                            formula_cols = [
                                'STT', 'Số ngày thi công', 'Tổng quỹ lương chi trả đến hiện tại của tổ', 
                                'Tổng sản lượng theo khối lượng', 'Chênh lệch Lãi(+)/Lỗ(-)', 
                                'ĐÁNH GIÁ & CẢNH BÁO'
                            ]
                            if col_name in formula_cols:
                                continue
                        elif mapped_sheet == 'ĐÁNH GIÁ TỔNG DA':
                            formula_cols = [
                                'STT', 'Số ngày thi công', 'Tổng quỹ lương chi trả đến hiện tại', 
                                'Tổng sản lượng theo khối lượng', 'Chênh lệch Lãi(+)/Lỗ(-)', 
                                'ĐÁNH GIÁ & CẢNH BÁO'
                            ]
                            if col_name in formula_cols:
                                continue
                                
                        val = record.get(col_name, '')
                        if val == '' or val is None:
                            ws.cell(row=r, column=col_idx, value=None)
                        else:
                            # Try parsing as date first
                            parsed_date = None
                            if isinstance(val, str) and ('-' in val or '/' in val):
                                import datetime
                                clean_val = val.split(' ')[0]
                                for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
                                    try:
                                        parsed_date = datetime.datetime.strptime(clean_val, fmt).date()
                                        break
                                    except ValueError:
                                        continue
                            
                            if parsed_date is not None:
                                ws.cell(row=r, column=col_idx, value=parsed_date)
                            else:
                                try:
                                    if '.' in str(val):
                                        val = float(val)
                                    else:
                                        val = int(val)
                                except ValueError:
                                    pass
                                ws.cell(row=r, column=col_idx, value=val)
                                
                    if mapped_sheet in ['GIAO VIỆC HÒA VỐN', 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY', 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG', 'ĐÁNH GIÁ TỔNG DA']:
                        apply_formulas_and_stt(ws)
                        
                    wb.save(excel_path)
                    self.send_json({"status": "success", "message": "Lưu bản ghi thành công!"})
                except PermissionError:
                    self.send_json({"status": "error", "message": "Không thể ghi tệp Excel. Vui lòng đóng tệp Excel trên máy tính của bạn trước khi lưu!"}, code=500)
                except Exception as e:
                    self.send_json({"status": "error", "message": f"Lỗi lưu file Excel: {str(e)}"}, code=500)

            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=400)
            return

        elif parsed.path == '/api/kpi/delete-record':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                sheet_name = data.get('sheet')
                row_index = int(data.get('rowIndex', -1))
                role = data.get('role', 'candidate')

                if not sheet_name or row_index < 0:
                    self.send_json({"status": "error", "message": "Thiếu dữ liệu sheet hoặc chỉ số dòng!"}, code=400)
                    return

                if sheet_name in ADMIN_ONLY_SHEETS and role != 'admin':
                    self.send_json({"status": "error", "message": "Bạn không có quyền xóa mục này! Cần quyền Admin."}, code=403)
                    return

                excel_path = get_cleaned_excel_path()
                
                # Delete from Google Sheets first
                gas_res = None
                try:
                    post_body = {
                        "action": "deleteRecord",
                        "sheet": sheet_name,
                        "rowIndex": row_index
                    }
                    gas_res = call_gas_api(post_data=post_body)
                    if gas_res.get('status') != 'success':
                        self.send_json(gas_res, code=400)
                        return
                except Exception as e:
                    self.send_json({"status": "error", "message": f"Lỗi đồng bộ Google Sheets: {str(e)}"}, code=500)
                    return
                    
                if not os.path.exists(excel_path):
                    self.send_json(gas_res)
                    return

                mapped_sheet = SHEETS_MAPPING.get(sheet_name)
                if not mapped_sheet:
                    self.send_json({"status": "error", "message": f"Sheet '{sheet_name}' không hợp lệ!"}, code=400)
                    return

                try:
                    wb = openpyxl.load_workbook(excel_path)
                    ws = wb[mapped_sheet]
                    
                    r = row_index + 2
                    if r < 2 or r > ws.max_row:
                        self.send_json({"status": "error", "message": "Chỉ số dòng không hợp lệ!"}, code=400)
                        return
                        
                    ws.delete_rows(r, 1)
                    
                    if mapped_sheet in ['GIAO VIỆC HÒA VỐN', 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY', 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG', 'ĐÁNH GIÁ TỔNG DA']:
                        apply_formulas_and_stt(ws)
                        
                    wb.save(excel_path)
                    self.send_json({"status": "success", "message": "Đã xóa bản ghi thành công!"})
                except PermissionError:
                    self.send_json({"status": "error", "message": "Không thể ghi tệp Excel. Vui lòng đóng tệp Excel trên máy tính của bạn trước khi lưu!"}, code=500)
                except Exception as e:
                    self.send_json({"status": "error", "message": f"Lỗi lưu file Excel: {str(e)}"}, code=500)

            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=400)
            return

        elif parsed.path == '/api/kpi/import-data':
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            try:
                data = json.loads(body.decode('utf-8'))
                sheet_name = data.get('sheet')
                imported_records = data.get('data')
                is_grid = data.get('isGrid', False)
                role = data.get('role', 'candidate')

                if not sheet_name or imported_records is None:
                    self.send_json({"status": "error", "message": "Thiếu dữ liệu sheet hoặc danh sách nhập!"}, code=400)
                    return

                if sheet_name in ADMIN_ONLY_SHEETS and role != 'admin':
                    self.send_json({"status": "error", "message": "Bạn không có quyền nhập dữ liệu cho mục này! Cần quyền Admin."}, code=403)
                    return

                excel_path = get_cleaned_excel_path()
                if not os.path.exists(excel_path):
                    try:
                        post_body = {
                            "action": "importData",
                            "sheet": sheet_name,
                            "isGrid": is_grid,
                            "data": imported_records
                        }
                        gas_res = call_gas_api(post_data=post_body)
                        if gas_res.get('status') == 'success':
                            self.send_json(gas_res)
                        else:
                            self.send_json(gas_res, code=400)
                    except Exception as e:
                        self.send_json({"status": "error", "message": f"Không tìm thấy tệp Excel cục bộ và lỗi kết nối Google Sheets: {str(e)}"}, code=500)
                    return

                mapped_sheet = SHEETS_MAPPING.get(sheet_name)
                if not mapped_sheet:
                    self.send_json({"status": "error", "message": f"Sheet '{sheet_name}' không hợp lệ!"}, code=400)
                    return

                try:
                    if is_grid:
                        # Ghi lưới trực tiếp
                        wb = openpyxl.load_workbook(excel_path)
                        if mapped_sheet in wb.sheetnames:
                            ws = wb[mapped_sheet]
                            ws.delete_rows(1, ws.max_row+1)
                        else:
                            ws = wb.create_sheet(title=mapped_sheet)
                        
                        for r_idx, row in enumerate(imported_records, 1):
                            for c_idx, val in enumerate(row, 1):
                                if val == "":
                                    ws.cell(row=r_idx, column=c_idx, value=None)
                                else:
                                    try:
                                        if '.' in str(val):
                                            ws.cell(row=r_idx, column=c_idx, value=float(val))
                                        else:
                                            ws.cell(row=r_idx, column=c_idx, value=int(val))
                                    except ValueError:
                                        ws.cell(row=r_idx, column=c_idx, value=str(val))
                        wb.save(excel_path)
                        self.send_json({"status": "success", "message": f"Nhập dữ liệu dạng lưới thành công cho {mapped_sheet}!"})
                    else:
                        df_new = pd.DataFrame(imported_records)
                        try:
                            df_old = pd.read_excel(excel_path, sheet_name=mapped_sheet)
                            for c in df_old.columns:
                                if c not in df_new.columns:
                                    df_new[c] = ''
                            df_new = df_new[df_old.columns]
                        except Exception:
                            pass
                            
                        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                            df_new.to_excel(writer, sheet_name=mapped_sheet, index=False)
                            
                        if mapped_sheet in ['GIAO VIỆC HÒA VỐN', 'BÁO CÁO SẢN LƯỢNG THỰC HIỆN TRONG NGÀY', 'ĐÁNH GIÁ TỔNG TRUNG ĐỘI TRƯỞNG', 'ĐÁNH GIÁ TỔNG DA']:
                            wb = openpyxl.load_workbook(excel_path)
                            ws = wb[mapped_sheet]
                            apply_formulas_and_stt(ws)
                            wb.save(excel_path)
                            
                        self.send_json({"status": "success", "message": f"Nhập dữ liệu thành công! Đã ghi nhận {len(df_new)} bản ghi mới."})
                except PermissionError:
                    self.send_json({"status": "error", "message": "Không thể ghi tệp Excel. Vui lòng đóng tệp Excel trên máy tính của bạn trước khi lưu!"}, code=500)
                except Exception as e:
                    self.send_json({"status": "error", "message": f"Lỗi lưu file Excel: {str(e)}"}, code=500)

            except Exception as e:
                self.send_json({"status": "error", "message": str(e)}, code=400)
            return

        self.send_error(404, "Endpoint not found")

    def do_DELETE(self):
        parsed = urlparse(self.path)
        query = parse_qs(parsed.query)
        role = query.get('role', [None])[0]
        
        if parsed.path == '/api/clear-results':
            if role != 'admin':
                self.send_json({"status": "error", "message": "Chỉ Giám thị mới có quyền xóa toàn bộ lịch sử thi!"}, code=403)
                return
                
            save_results_to_files([])
            self.send_json({"status": "success", "message": "Đã xóa toàn bộ lịch sử bài thi"})
            return
            
        elif parsed.path == '/api/delete-result':
            if role != 'admin':
                self.send_json({"status": "error", "message": "Chỉ Giám thị mới có quyền xóa kết quả bài thi!"}, code=403)
                return
                
            item_id = query.get('id', [None])[0]
            if item_id:
                results = load_results()
                filtered = [r for r in results if r.get('id') != item_id]
                save_results_to_files(filtered)
                self.send_json({"status": "success", "message": f"Đã xóa bản ghi {item_id}"})
            else:
                self.send_json({"status": "error", "message": "Missing id parameter"}, code=400)
            return

        self.send_error(404, "Endpoint not found")

def open_browser():
    webbrowser.open(f"http://localhost:{PORT}/index.html")

def run_server():
    socketserver.TCPServer.allow_reuse_address = True
    try:
        httpd = socketserver.TCPServer(("", PORT), Handler)
    except OSError as e:
        # Nếu cổng 8000 đã có tiến trình khác chiếm dụng (máy chủ đang chạy nền)
        # Chúng ta chỉ cần mở trình duyệt và thoát nhẹ nhàng
        print(f"\n[THÔNG BÁO] Máy chủ Vincons đang chạy sẵn ở cổng {PORT} trong nền.")
        print("Đang mở trình duyệt giao diện trên máy tính của bạn...")
        open_browser()
        return

    with httpd:
        print(f"Vincons Recruitment Exam Server running at http://localhost:{PORT}")
        print("Press Ctrl+C to stop.")
        threading.Timer(0.5, open_browser).start()
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\nStopping server...")
            httpd.shutdown()

if __name__ == '__main__':
    run_server()
